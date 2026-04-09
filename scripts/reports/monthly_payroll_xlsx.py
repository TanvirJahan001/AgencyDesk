"""
Monthly Payroll Report — Excel Export

Reads JSON from stdin:
{
  "companyName": "...",
  "period": "2026-04",
  "periodStart": "2026-04-01",
  "periodEnd": "2026-04-30",
  "generatedAt": "...",
  "runs": [
    {
      "employeeName": "...",
      "department": "...",
      "hourlyRate": 25.0,
      "overtimeMultiplier": 1.5,
      "totalWorkHours": 45.5,
      "regularHours": 40.0,
      "overtimeHours": 5.5,
      "regularPay": 1000.0,
      "overtimePay": 206.25,
      "grossPay": 1206.25,
      "deductions": 50.0,
      "netPay": 1156.25,
      "status": "processed"
    }
  ],
  "totals": {
    "totalWorkHours": 91.0,
    "regularHours": 80.0,
    "overtimeHours": 11.0,
    "regularPay": 2000.0,
    "overtimePay": 412.50,
    "grossPay": 2412.50,
    "deductions": 100.0,
    "netPay": 2312.50
  }
}

Writes .xlsx to argv[1].
"""

import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

data = json.load(sys.stdin)
out_path = sys.argv[1]

wb = Workbook()
ws = wb.active
ws.title = "Payroll Summary"

BRAND = "1E40AF"
HDR_FILL = PatternFill("solid", fgColor=BRAND)
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color=BRAND)
SUB_FONT = Font(name="Arial", size=10, color="666666")
DATA = Font(name="Arial", size=10)
BOLD = Font(name="Arial", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="EBF5FB")
PAID_FILL = PatternFill("solid", fgColor="E6F9E6")
BORDER = Border(
    left=Side("thin", color="D0D5DD"), right=Side("thin", color="D0D5DD"),
    top=Side("thin", color="D0D5DD"), bottom=Side("thin", color="D0D5DD"),
)
C = Alignment(horizontal="center", vertical="center")
R = Alignment(horizontal="right", vertical="center")
L = Alignment(horizontal="left", vertical="center")
MONEY_FMT = '$#,##0.00'

# --- Title ---
ws.merge_cells("A1:K1")
ws["A1"] = data.get("companyName", "AgencyDesk")
ws["A1"].font = TITLE_FONT

ws.merge_cells("A2:K2")
ws["A2"] = f"Payroll Report — {data['period']}  ({data['periodStart']} to {data['periodEnd']})"
ws["A2"].font = SUB_FONT

ws.merge_cells("A3:K3")
ws["A3"] = f"Generated: {data.get('generatedAt', '')}"
ws["A3"].font = Font(name="Arial", size=9, color="999999")

# --- Headers ---
row = 5
headers = [
    "Employee", "Dept", "Rate ($/hr)", "Total Hrs",
    "Regular Hrs", "OT Hrs", "Regular Pay", "OT Pay",
    "Gross Pay", "Deductions", "Net Pay"
]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=c, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = C
    cell.border = BORDER
row += 1

# --- Data ---
money_cols = {3, 7, 8, 9, 10, 11}
hr_cols = {4, 5, 6}

for run in data.get("runs", []):
    vals = [
        run["employeeName"],
        run.get("department", "—"),
        run["hourlyRate"],
        round(run["totalWorkHours"], 2),
        round(run["regularHours"], 2),
        round(run["overtimeHours"], 2),
        round(run["regularPay"], 2),
        round(run["overtimePay"], 2),
        round(run["grossPay"], 2),
        round(run["deductions"], 2),
        round(run["netPay"], 2),
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = DATA
        cell.border = BORDER
        if c in money_cols:
            cell.number_format = MONEY_FMT
            cell.alignment = R
        elif c in hr_cols:
            cell.number_format = "0.00"
            cell.alignment = R
        else:
            cell.alignment = L

        if run.get("status") == "paid":
            cell.fill = PAID_FILL
    row += 1

# --- Totals row ---
totals = data.get("totals", {})
total_vals = [
    "TOTALS", "",
    "",
    round(totals.get("totalWorkHours", 0), 2),
    round(totals.get("regularHours", 0), 2),
    round(totals.get("overtimeHours", 0), 2),
    round(totals.get("regularPay", 0), 2),
    round(totals.get("overtimePay", 0), 2),
    round(totals.get("grossPay", 0), 2),
    round(totals.get("deductions", 0), 2),
    round(totals.get("netPay", 0), 2),
]
for c, v in enumerate(total_vals, 1):
    cell = ws.cell(row=row, column=c, value=v)
    cell.font = BOLD
    cell.fill = TOTAL_FILL
    cell.border = BORDER
    if c in money_cols:
        cell.number_format = MONEY_FMT
        cell.alignment = R
    elif c in hr_cols:
        cell.number_format = "0.00"
        cell.alignment = R
    else:
        cell.alignment = L

# --- Column widths ---
col_widths = [22, 14, 12, 12, 12, 10, 14, 14, 14, 14, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.print_title_rows = "1:5"

wb.save(out_path)
print(json.dumps({"status": "success", "path": out_path}))
