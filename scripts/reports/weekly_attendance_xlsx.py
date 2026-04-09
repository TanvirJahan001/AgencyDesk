"""
Weekly Attendance Report — Excel Export

Reads JSON from stdin with shape:
{
  "companyName": "...",
  "periodLabel": "2026-W15",
  "periodStart": "2026-04-06",
  "periodEnd": "2026-04-12",
  "generatedAt": "...",
  "employees": [
    {
      "name": "...",
      "department": "...",
      "days": [
        { "date": "2026-04-06", "workHours": 8.0, "breakHours": 0.5, "status": "completed" },
        ...
      ],
      "totalWorkHours": 40.0,
      "totalBreakHours": 2.5,
      "daysWorked": 5
    }
  ]
}

Writes .xlsx to the path given as argv[1].
"""

import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.load(sys.stdin)
out_path = sys.argv[1]

wb = Workbook()
ws = wb.active
ws.title = "Weekly Attendance"

# --- Styles ---
BRAND_BLUE = "1E40AF"
HEADER_FILL = PatternFill("solid", fgColor=BRAND_BLUE)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color=BRAND_BLUE)
SUBTITLE_FONT = Font(name="Arial", size=10, color="666666")
DATA_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="EBF5FB")
ABSENT_FILL = PatternFill("solid", fgColor="FEF3F2")
BORDER_THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

# --- Header ---
ws.merge_cells("A1:I1")
ws["A1"] = data.get("companyName", "AgencyDesk")
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = LEFT

ws.merge_cells("A2:I2")
ws["A2"] = f"Weekly Attendance Report — {data['periodLabel']}  ({data['periodStart']} to {data['periodEnd']})"
ws["A2"].font = SUBTITLE_FONT

ws.merge_cells("A3:I3")
ws["A3"] = f"Generated: {data.get('generatedAt', '')}"
ws["A3"].font = Font(name="Arial", size=9, color="999999")

row = 5

# Day labels
days_of_week = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

for emp in data["employees"]:
    # Employee header
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=f"{emp['name']}  —  {emp.get('department', 'N/A')}")
    cell.font = Font(name="Arial", bold=True, size=11, color=BRAND_BLUE)
    cell.fill = PatternFill("solid", fgColor="F0F4FF")
    cell.border = BORDER_THIN
    row += 1

    # Column headers
    headers = ["Day", "Date", "Work (hrs)", "Break (hrs)", "Status"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    row += 1

    # Day rows
    for i, day in enumerate(emp.get("days", [])):
        day_name = days_of_week[i] if i < 7 else ""
        is_absent = day.get("status") == "absent"

        ws.cell(row=row, column=1, value=day_name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=day["date"]).font = DATA_FONT
        ws.cell(row=row, column=3, value=round(day.get("workHours", 0), 2) if not is_absent else 0).font = DATA_FONT
        ws.cell(row=row, column=4, value=round(day.get("breakHours", 0), 2) if not is_absent else 0).font = DATA_FONT
        ws.cell(row=row, column=5, value="Absent" if is_absent else "Present").font = DATA_FONT

        for c in range(1, 6):
            ws.cell(row=row, column=c).border = BORDER_THIN
            ws.cell(row=row, column=c).alignment = CENTER if c >= 3 else LEFT
            if is_absent:
                ws.cell(row=row, column=c).fill = ABSENT_FILL
        row += 1

    # Summary row
    summary_labels = ["", "TOTAL", emp["totalWorkHours"], emp["totalBreakHours"], f"{emp['daysWorked']} days"]
    for c, val in enumerate(summary_labels, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER_THIN
        cell.alignment = CENTER if c >= 3 else LEFT
    row += 2

# --- Column widths ---
widths = [8, 14, 14, 14, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Print setup ---
ws.print_title_rows = "1:4"
ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr or type(ws.sheet_properties.pageSetUpPr)()

wb.save(out_path)
print(json.dumps({"status": "success", "path": out_path}))
